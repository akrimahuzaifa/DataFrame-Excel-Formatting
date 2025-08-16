from openpyxl.styles import PatternFill, Font, Alignment

def set_cell_color(sheet, cell, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    sheet[cell].fill = fill

def set_font_style(sheet, cell, bold=False, italic=False):
    font = Font(bold=bold, italic=italic)
    sheet[cell].font = font

def set_alignment(sheet, cell, horizontal='center', vertical='center'):
    alignment = Alignment(horizontal=horizontal, vertical=vertical)
    sheet[cell].alignment = alignment

def format_currency(sheet, cell):
    sheet[cell].number_format = '"$"#,##0.00'

def format_percentage(sheet, cell):
    sheet[cell].number_format = '0.00%'

def merge_and_center(sheet, start_cell, end_cell):
    sheet.merge_cells(start_row=start_cell[0], start_column=start_cell[1],
                      end_row=end_cell[0], end_column=end_cell[1])
    merged_cell = sheet.cell(row=start_cell[0], column=start_cell[1])
    set_alignment(sheet, merged_cell.coordinate, horizontal='center', vertical='center')