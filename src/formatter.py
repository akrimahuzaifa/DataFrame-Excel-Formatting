from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def create_excel_report(data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Set header
    headers = data[0].keys()
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Fill data
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data.values(), 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if isinstance(value, (int, float)):
                if "amount" in row_data:  # Assuming 'amount' indicates a dollar value
                    cell.number_format = '"$"#,##0.00'
                elif "percentage" in row_data:  # Assuming 'percentage' indicates a percentage value
                    cell.number_format = '0.00%'
            # Example of filling a cell with color based on some condition
            if value < 0:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red fill

    # Merge and center rows if needed
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    merged_cell = ws.cell(row=1, column=1)
    merged_cell.value = "Formatted Report"
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_file)