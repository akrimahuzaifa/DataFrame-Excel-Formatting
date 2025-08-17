from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from ..utils.styles import *

def create_example_template():
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Example Report"

    # Set header values
    headers = ["Item", "Amount", "Percentage"]
    ws.append(headers)

    # Style the header
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        set_font_style(ws, cell.coordinate, bold=True)
        set_alignment(ws, cell.coordinate, horizontal='center')
        set_cell_color(ws, cell.coordinate, color="00FFCC")  # Light green background

    # Add data
    data = [
        ["Item A", 1000, 0.25],
        ["Item B", 1500, 0.50],
        ["Item C", 2000, 0.75],
    ]

    for item, amount, percentage in data:
        ws.append([item, amount, percentage])

    # Format Amount and Percentage columns
    for row in range(2, len(data) + 2):
        ws.cell(row=row, column=2).number_format = '"$"#,##0.00'  # Dollar format
        ws.cell(row=row, column=3).number_format = '0.00%'  # Percentage format

    # Merge and center a title row
    ws.merge_cells('A1:C1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Sales Report"
    set_font_style(ws, title_cell.coordinate, bold=True)
    set_alignment(ws, title_cell.coordinate, horizontal='center', vertical='center')

    # Apply borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            set_border(ws, cell.coordinate)

    # Save the workbook
    wb.save("example_report.xlsx")

if __name__ == "__main__":
    create_example_template()